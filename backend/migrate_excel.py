import openpyxl
import sqlite3
import os

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"
db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "collection.db")

def main():
    if not os.path.exists(excel_path):
        print(f"Excel file not found at {excel_path}")
        return

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS collection (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            brand TEXT,
            name TEXT,
            season TEXT,
            strength TEXT,
            rating TEXT,
            longevity TEXT,
            projection TEXT,
            price_tier TEXT,
            inspiration TEXT
        )
    ''')
    conn.commit()
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Failed to open excel: {e}")
        return
        
    if "Collection" not in wb.sheetnames:
        print("Collection sheet not found.")
        return
        
    sheet = wb["Collection"]
    
    count = 0
    # Clear existing to prevent duplicates during testing
    cursor.execute("DELETE FROM collection")
    
    for r in range(2, sheet.max_row + 1):
        brand = sheet.cell(row=r, column=2).value
        name = sheet.cell(row=r, column=3).value
        if brand and name:
            season = sheet.cell(row=r, column=4).value or "All Season"
            strength = sheet.cell(row=r, column=5).value or "Moderate"
            rating = str(sheet.cell(row=r, column=6).value or "8.0")
            longevity = sheet.cell(row=r, column=7).value or "Moderate"
            projection = sheet.cell(row=r, column=8).value or "Moderate"
            price = sheet.cell(row=r, column=9).value or "Budget"
            
            try:
                inspiration = sheet.cell(row=r, column=13).value or "Original"
            except:
                inspiration = "Original"
            
            cursor.execute('''
                INSERT INTO collection (brand, name, season, strength, rating, longevity, projection, price_tier, inspiration)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (brand, name, season, strength, rating, longevity, projection, price, inspiration))
            count += 1
            
    conn.commit()
    conn.close()
    print(f"Successfully migrated {count} fragrances from Excel to SQLite!")

if __name__ == "__main__":
    main()
