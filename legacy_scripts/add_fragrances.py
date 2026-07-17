import openpyxl
import sys
import shutil
from pathlib import Path

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"
backup_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker_backup.xlsx"
shutil.copy2(excel_path, backup_path)

data = [
    ("Nautica", "Voyage"),
    ("Bujairami", "Ruthless Vanilla"),
    ("Amaran Parfums", "Vault Men"),
    ("Miniso's", "Black Gravitation"),
    ("Maison Alhambra", "Victorioso Victory"),
    ("Divain's", "376 (Imagination)"),
    ("Divain's", "358 (Meteore)"),
    ("Divain's", "737 (Sicily)"),
    ("Divain's", "677 (Afternoon Swim)"),
    ("Divain's", "230 (One Million Prive)"),
    ("Divain's", "240 (Wanted)"),
    ("Divain's", "330 (Red Tabacco)"),
    ("Lattafa", "Qaa'ed"),
    ("Dua's", "Magnetic Scent For Him"),
    ("KDJ", "Prada Luna Rossa Ocean"),
    ("Chez Pierre", "Imperialism"),
    ("Chez Pierre", "Warrior"),
    ("Chez Pierre", "Secret Garden"),
    ("Chez Pierre", "Tuxedo"),
    ("Chez Pierre", "Adventure"),
    ("Armaf", "The Lions Club Féroce"),
    ("Armaf", "The Lions Club Monarque"),
    ("Rabanne", "Paco Edt"),
    ("Armaf", "Craze"),
    ("FAAN", "Momento Matrix"),
    ("Niche Emarati Parfums", "Al Jawhara"),
    ("Grandeur Elite", "Tribal Gold"),
    ("Lucianno", "Dynasty"),
    ("Sergio Vallanti", "Tresador Mystique"),
    ("Paris Corner", "Rifaaqat"),
    ("Bruce Buffer", "It's Time"),
    ("Pendora Scents", "Milano Privé"),
    ("Vurv", "Furtune Lucky Man"),
    ("FAAN", "Mashair Indigo"),
    ("Ainash", "Royal Noir Absolu"),
    ("Fragrance World", "F Le Parfum"),
    ("Emir", "Factory Campfire"),
    ("Karisma", "Apple Orchard"),
    ("Dossier", "Citrus Ginger"),
    ("Metropolis", "9th Avenue New York"),
    ("Arabiyat Prestige", "Bamboo Reflection"),
    ("Zakat", "You Are My Fire Rouge"),
    ("Lattafa", "Mashrabya"),
    ("Fragrance World", "Nazih Gold"),
    ("Armaf", "Dunescape"),
    ("Ahmed", "Kaaf Noir"),
    ("Armaf", "Tres Nuit Lyric"),
    ("Armaf", "Tres Nuit"),
    ("Macarena", "Summer Swim"),
    ("Macarena", "MegaMarine"),
    ("Macarena", "Symphony 33"),
    ("Armaf", "The Pride Of Armaf"),
    ("Armaf", "Shades Blue"),
    ("Maison Alhambra", "Luxe Bold"),
    ("Maison Alhambra", "Luxe Journey"),
    ("Maison Alhambra", "Luxe Gold"),
    ("Zakat", "Z25"),
    ("Zakat", "Z9"),
    ("Lattafa", "Sehr"),
    ("Legend", "Legend Pour Homme"),
    ("Maycreate", "Gather Beauty"),
    ("Armaf", "Excellus"),
    ("Fragrance World", "Mocha Wood"),
    ("Milestone", "Red Cherry"),
    ("Azzurra Parfums", "Golden Amber"),
    ("Fragrance World", "Al Raiee"),
    ("Armaf", "Dubai Nights Midnight"),
    ("Grandeur Elite", "Calabria Cerulean"),
    ("Armaf", "Club De Nuit Intense Edt"),
    ("Armaf", "Club De Nuit Intense Edp"),
    ("Armaf", "Club De Nuit Limited Edition Parfum"),
    ("Armaf", "Club De Nuit Urban Man Elixir"),
    ("Armaf", "Club De Nuit Sillage"),
    ("Armaf", "Club De Nuit Milestone"),
    ("Armaf", "Club De Nuit Bling"),
    ("Armaf", "Club De Nuit Precieux"),
    ("Armaf", "Club De Nuit Private Key To My Dreams"),
    ("Armaf", "Club De Nuit Private Key To My Success"),
    ("Armaf", "Club De Nuit Private Key To My Life"),
    ("Armaf", "Club De Nuit Private Key To My Love"),
]

wb = openpyxl.load_workbook(excel_path)
sheet = wb["Collection"]

last_row = 2
for r in range(2, sheet.max_row + 1):
    if sheet.cell(row=r, column=2).value or sheet.cell(row=r, column=3).value:
        last_row = r

start_num = 1
val = sheet.cell(row=last_row, column=1).value
if isinstance(val, (int, float)):
    start_num = int(val) + 1
else:
    cnt = 0
    for r in range(2, last_row + 1):
        if sheet.cell(row=r, column=2).value:
            cnt += 1
    start_num = cnt + 1

current_row = last_row + 1
for brand, name in data:
    sheet.cell(row=current_row, column=1, value=start_num)
    sheet.cell(row=current_row, column=2, value=brand)
    sheet.cell(row=current_row, column=3, value=name)
    start_num += 1
    current_row += 1

wb.save(excel_path)
print(f"Successfully appended {len(data)} fragrances.")
