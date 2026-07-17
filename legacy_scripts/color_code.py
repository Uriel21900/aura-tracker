import openpyxl
from openpyxl.styles import PatternFill

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"

# ARGB Hex colors for conditional formatting
season_colors = {
    "Spring": "FFD9EAD3",        # Light Green
    "Summer": "FFFFE599",        # Light Yellow
    "Fall": "FFFCE5CD",          # Light Orange
    "Winter": "FFCFE2F3",        # Light Blue
    "All Season": "FFD9D2E9",    # Light Purple
    "Spring/Summer": "FFD9EAD3", # Light Green
}

longevity_colors = {
    "Weak": "FFF4CCCC",          # Light Red
    "Moderate": "FFFFE599",      # Light Yellow
    "Long": "FFD9EAD3",          # Light Green
    "Long Lasting": "FFD9EAD3",  # Light Green
    "Eternal": "FF93C47D",       # Medium Green
}

try:
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb["Collection"]
    
    count = 0
    for r in range(3, sheet.max_row + 1):
        # Season is column 4 (D)
        season_val = sheet.cell(row=r, column=4).value
        if season_val and str(season_val).strip() in season_colors:
            fill_color = season_colors[str(season_val).strip()]
            sheet.cell(row=r, column=4).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            count += 1
            
        # Longevity is column 7 (G)
        longevity_val = sheet.cell(row=r, column=7).value
        if longevity_val and str(longevity_val).strip() in longevity_colors:
            fill_color = longevity_colors[str(longevity_val).strip()]
            sheet.cell(row=r, column=7).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            count += 1
            
    wb.save(excel_path)
    print(f"Applied color coding to {count} cells successfully.")
except Exception as e:
    print(f"Error: {e}")
