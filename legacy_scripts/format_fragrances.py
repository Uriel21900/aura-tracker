import openpyxl
from copy import copy

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb["Collection"]

    # We will copy the formatting from row 3 (the first "old" fragrance)
    source_row = 3
    
    count = 0
    # Apply to all rows that have data, skipping the source row itself
    for r in range(4, sheet.max_row + 1):
        if sheet.cell(row=r, column=2).value or sheet.cell(row=r, column=3).value:
            for col in range(1, 13):
                source_cell = sheet.cell(row=source_row, column=col)
                target_cell = sheet.cell(row=r, column=col)
                
                # Copy border and other styling properties
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
            count += 1

    wb.save(excel_path)
    print(f"Applied formatting to {count} rows successfully.")
except Exception as e:
    print(f"Error: {e}")
