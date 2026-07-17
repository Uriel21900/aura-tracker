import openpyxl
from copy import copy

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"

dupe_mapping = {
    "Voyage": "Original",
    "Power by 50 Cent": "Original",
    "Maximum": "Original",
    "9 PM Night Out": "Dupe of JPG Ultra Male",
    "9pm": "Dupe of JPG Ultra Male",
    "Historic Olmeda": "Dupe of Bleu de Chanel",
    "Supremacy Collector's Edition Pour Homme": "Dupe of Creed Aventus",
    "Supremacy Not Only Intense": "Dupe of Nishane Hacivat / Aventus",
    "Ruthless Vanilla": "Dupe of YSL Babycat / Vanilla",
    "Vault Men": "Dupe of Armani Code Profumo",
    "Black Gravitation": "Smells like Dior Sauvage",
    "Victorioso Victory": "Dupe of Paco Rabanne Invictus Victory",
    "376 (Imagination)": "Dupe of Louis Vuitton Imagination",
    "358 (Meteore)": "Dupe of Louis Vuitton Météore",
    "737 (Sicily)": "Dupe of Mancera Sicily",
    "677 (Afternoon Swim)": "Dupe of Louis Vuitton Afternoon Swim",
    "230 (One Million Prive)": "Dupe of Paco Rabanne 1 Million Privé",
    "240 (Wanted)": "Dupe of Azzaro Wanted",
    "330 (Red Tabacco)": "Dupe of Mancera Red Tobacco",
    "Qaa'ed": "Original (Smells like Dunhill Icon Absolute)",
    "Magnetic Scent For Him": "Dupe of Escada Magnetism for Men",
    "Prada Luna Rossa Ocean": "Dupe of Prada Luna Rossa Ocean",
    "Imperialism": "Dupe of Creed Millésime Impérial",
    "Warrior": "Dupe of Creed Viking",
    "Secret Garden": "Dupe of Tom Ford Noir de Noir",
    "Tuxedo": "Dupe of YSL Tuxedo",
    "Adventure": "Dupe of Creed Aventus",
    "The Lions Club Féroce": "Smells like Dior Sauvage / Fierce",
    "The Lions Club Monarque": "Smells like Niche Oriental",
    "Paco Edt": "Original",
    "Craze": "Dupe of Parfums de Marly Pegasus",
    "Momento Matrix": "Smells like Designer Fresh",
    "Al Jawhara": "Original",
    "Tribal Gold": "Smells like Sweet Niche",
    "Dynasty": "Dupe of PDM Haltane / Oud for Greatness",
    "Tresador Mystique": "Dupe of Lancome Tresor Midnight Rose",
    "Rifaaqat": "Dupe of YSL Babycat",
    "It's Time": "Original",
    "Milano Privé": "Dupe of Paco Rabanne 1 Million Privé",
    "Furtune Lucky Man": "Dupe of Paco Rabanne 1 Million Lucky",
    "Mashair Indigo": "Dupe of Dior Sauvage",
    "Royal Noir Absolu": "Dupe of Tom Ford Noir Extreme",
    "F Le Parfum": "Dupe of YSL Y Le Parfum",
    "Factory Campfire": "Dupe of Maison Margiela By the Fireplace",
    "Apple Orchard": "Dupe of Kilian Apple Brandy",
    "Citrus Ginger": "Dupe of Bleu de Chanel",
    "9th Avenue New York": "Dupe of Bond No 9",
    "Bamboo Reflection": "Dupe of Amouage Reflection Man",
    "You Are My Fire Rouge": "Dupe of MFK Baccarat Rouge 540",
    "Mashrabya": "Original",
    "Nazih Gold": "Smells like Oriental Sweet",
    "Dunescape": "Dupe of Maison Margiela Beach Walk",
    "Kaaf Noir": "Dupe of Abercrombie & Fitch Fierce",
    "Tres Nuit Lyric": "Dupe of Armani Acqua di Gio Profondo",
    "Tres Nuit": "Dupe of Creed Green Irish Tweed",
    "Summer Swim": "Dupe of Louis Vuitton Afternoon Swim",
    "MegaMarine": "Dupe of Orto Parisi Megamare",
    "Symphony 33": "Dupe of Le Labo Santal 33",
    "The Pride Of Armaf": "Dupe of Dior Sauvage",
    "Shades Blue": "Dupe of Bleu de Chanel",
    "Luxe Bold": "Smells like Rich Oriental",
    "Luxe Journey": "Smells like Fresh Oriental",
    "Luxe Gold": "Smells like Sweet Amber",
    "Z25": "Dupe of Le Labo Tonka 25",
    "Z9": "Smells like Niche",
    "Sehr": "Original",
    "Legend Pour Homme": "Dupe of Montblanc Legend",
    "Gather Beauty": "Smells like Fresh Designer",
    "Excellus": "Smells like Designer Fresh",
    "Mocha Wood": "Dupe of Franck Boclet Tobacco",
    "Red Cherry": "Dupe of Tom Ford Lost Cherry",
    "Golden Amber": "Smells like Amber Niche",
    "Al Raiee": "Dupe of D&G The One Luminous Night",
    "Dubai Nights Midnight": "Smells like Rich Oud",
    "Calabria Cerulean": "Dupe of Acqua di Parma Mirto di Panarea",
    "Club De Nuit Intense Edt": "Dupe of Creed Aventus",
    "Club De Nuit Intense Edp": "Dupe of Creed Aventus",
    "Club De Nuit Limited Edition Parfum": "Dupe of Creed Aventus",
    "Club De Nuit Urban Man Elixir": "Dupe of Dior Sauvage / Aventus",
    "Club De Nuit Sillage": "Dupe of Creed Silver Mountain Water",
    "Club De Nuit Milestone": "Dupe of Creed Millésime Impérial",
    "Club De Nuit Bling": "Smells like Chanel Coco Mademoiselle",
    "Club De Nuit Precieux": "Dupe of Creed Aventus Absolu",
    "Club De Nuit Private Key To My Dreams": "Original",
    "Club De Nuit Private Key To My Success": "Original",
    "Club De Nuit Private Key To My Life": "Original",
    "Club De Nuit Private Key To My Love": "Original",
}

try:
    wb = openpyxl.load_workbook(excel_path)
    
    tabs = ["Collection", "Sorted by Season", "Sorted by Strength"]
    for tab in tabs:
        if tab in wb.sheetnames:
            sheet = wb[tab]
            
            # Setup Header for Column 13 (M)
            header_cell = sheet.cell(row=2, column=13)
            header_cell.value = "Inspiration / Type"
            # Copy style from column 12 header
            source_header = sheet.cell(row=2, column=12)
            if source_header.has_style:
                header_cell.font = copy(source_header.font)
                header_cell.border = copy(source_header.border)
                header_cell.fill = copy(source_header.fill)
                header_cell.alignment = copy(source_header.alignment)
            
            for r in range(3, sheet.max_row + 1):
                name = sheet.cell(row=r, column=3).value
                if name:
                    name_str = str(name).strip()
                    val = dupe_mapping.get(name_str, "Original / Unknown")
                    
                    target_cell = sheet.cell(row=r, column=13)
                    target_cell.value = val
                    
                    # Copy styling from column 12 (Times Worn)
                    source_cell = sheet.cell(row=r, column=12)
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.alignment = copy(source_cell.alignment)

    # Automatically adjust width of column M
    for tab in tabs:
        if tab in wb.sheetnames:
            sheet = wb[tab]
            max_length = 0
            for cell in sheet['M']:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions['M'].width = adjusted_width

    wb.save(excel_path)
    print("Successfully added the Inspiration/Type column.")
except Exception as e:
    print(f"Error: {e}")
