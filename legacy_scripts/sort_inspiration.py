import openpyxl
from copy import copy

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    source_ws = wb["Collection"]

    tab_title = "Sorted by Inspiration"
    
    # Delete if it already exists
    if tab_title in wb.sheetnames:
        del wb[tab_title]
        
    ws_new = wb.copy_worksheet(source_ws)
    ws_new.title = tab_title

    # Extract all rows from 3 to max_row
    rows_data = []
    for r in range(3, ws_new.max_row + 1):
        # Ensure row has data in Brand or Name
        if ws_new.cell(row=r, column=2).value or ws_new.cell(row=r, column=3).value:
            row_cells = []
            for c in range(1, 14): # Up to 13 (Column M)
                cell = ws_new.cell(row=r, column=c)
                cell_data = {
                    'value': cell.value,
                    'data_type': cell.data_type,
                    'font': copy(cell.font) if cell.has_style else None,
                    'border': copy(cell.border) if cell.has_style else None,
                    'fill': copy(cell.fill) if cell.has_style else None,
                    'number_format': copy(cell.number_format) if cell.has_style else None,
                    'protection': copy(cell.protection) if cell.has_style else None,
                    'alignment': copy(cell.alignment) if cell.has_style else None,
                }
                row_cells.append(cell_data)
            rows_data.append(row_cells)
    
    # Sort based on Column M (Index 12)
    def get_sort_key(row):
        val = row[12]['value'] # Column 13 (M)
        return str(val).lower() if val is not None else ""
        
    rows_data.sort(key=get_sort_key)
    
    # Write sorted data back to the sheet
    for i, row_cells in enumerate(rows_data):
        r = i + 3
        for j, cell_data in enumerate(row_cells):
            c = j + 1
            cell = ws_new.cell(row=r, column=c)
            
            # Reconstruct formulas and sequential IDs
            if c == 1:
                cell.value = i + 1
            elif c == 12: # Times Worn Formula
                cell.value = f"=COUNTIF('Wear Log'!B:B,C{r})"
            elif c == 11: # Last Worn Formula
                cell.value = f'=IF(COUNTIF(\'Wear Log\'!B:B,C{r})>0, MAXIFS(\'Wear Log\'!A:A, \'Wear Log\'!B:B, C{r}), "")'
            else:
                cell.value = cell_data['value']
                cell.data_type = cell_data['data_type']
                
            # Keep styles
            if cell_data['font']: cell.font = cell_data['font']
            if cell_data['border']: cell.border = cell_data['border']
            if cell_data['fill']: cell.fill = cell_data['fill']
            if cell_data['number_format']: cell.number_format = cell_data['number_format']
            if cell_data['protection']: cell.protection = cell_data['protection']
            if cell_data['alignment']: cell.alignment = cell_data['alignment']

    wb.save(excel_path)
    print("Successfully created 'Sorted by Inspiration' tab.")

except Exception as e:
    print(f"Error: {e}")
