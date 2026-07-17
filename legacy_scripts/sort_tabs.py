import openpyxl
from copy import copy

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    source_ws = wb["Collection"]

    # Create new tabs
    if "Sorted by Season" in wb.sheetnames:
        del wb["Sorted by Season"]
    ws_season = wb.copy_worksheet(source_ws)
    ws_season.title = "Sorted by Season"

    if "Sorted by Strength" in wb.sheetnames:
        del wb["Sorted by Strength"]
    ws_strength = wb.copy_worksheet(source_ws)
    ws_strength.title = "Sorted by Strength"

    def sort_sheet(ws, sort_col_idx):
        rows_data = []
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=2).value or ws.cell(row=r, column=3).value:
                row_cells = []
                for c in range(1, 13):
                    cell = ws.cell(row=r, column=c)
                    cell_data = {
                        'value': cell.value,
                        'data_type': cell.data_type,
                        'font': copy(cell.font) if cell.has_style else None,
                        'border': copy(cell.border) if cell.has_style else None,
                        'fill': copy(cell.fill) if cell.has_style else None,
                        'number_format': copy(cell.number_format) if cell.has_style else None,
                        'protection': copy(cell.protection) if cell.has_style else None,
                        'alignment': copy(cell.alignment) if cell.has_style else None,
                    }
                    row_cells.append(cell_data)
                rows_data.append(row_cells)
        
        # Sort rows based on the requested column index (0-indexed in row_cells)
        def get_sort_key(row):
            val = row[sort_col_idx]['value']
            return str(val) if val is not None else ""
            
        rows_data.sort(key=get_sort_key)
        
        # Write the sorted data back to the worksheet
        for i, row_cells in enumerate(rows_data):
            r = i + 3
            for j, cell_data in enumerate(row_cells):
                c = j + 1
                cell = ws.cell(row=r, column=c)
                # Overwrite value except for formulas that depend on row number
                if c == 12: # Times Worn formula
                    cell.value = f"=COUNTIF('Wear Log'!B:B,C{r})"
                elif c == 1: # Regenerate the sequential ID number
                    cell.value = i + 1
                else:
                    cell.value = cell_data['value']
                    cell.data_type = cell_data['data_type']
                
                # Apply styles
                if cell_data['font']: cell.font = cell_data['font']
                if cell_data['border']: cell.border = cell_data['border']
                if cell_data['fill']: cell.fill = cell_data['fill']
                if cell_data['number_format']: cell.number_format = cell_data['number_format']
                if cell_data['protection']: cell.protection = cell_data['protection']
                if cell_data['alignment']: cell.alignment = cell_data['alignment']

    # Season is Column 4 (index 3)
    sort_sheet(ws_season, 3)

    # Strength is Column 5 (index 4)
    sort_sheet(ws_strength, 4)

    # Reorder the sheets so the original Collection is first
    # 'Collection' is usually first anyway, but `copy_worksheet` puts new sheets at the end
    
    wb.save(excel_path)
    print("Successfully created tabs and sorted them.")

except Exception as e:
    print(f"Error: {e}")
