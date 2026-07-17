import openpyxl
import csv
import os

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"
dataset_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dataset")

def load_dataset():
    """
    Looks for the first CSV file in the dataset folder and loads it into a dictionary for quick lookup.
    """
    database = {}
    if not os.path.exists(dataset_path):
        return database
        
    csv_files = [f for f in os.listdir(dataset_path) if f.endswith('.csv')]
    if not csv_files:
        return database
        
    csv_file = os.path.join(dataset_path, csv_files[0])
    
    try:
        with open(csv_file, mode='r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            # We will map these to the actual column names in the dataset once downloaded.
            for row in reader:
                # Find columns that match "brand" or "name" regardless of capitalization
                brand_col = next((c for c in row.keys() if c and 'brand' in c.lower()), None)
                name_col = next((c for c in row.keys() if c and ('name' in c.lower() or 'perfume' in c.lower())), None)
                
                if brand_col and name_col:
                    brand = str(row[brand_col]).strip().lower()
                    name = str(row[name_col]).strip().lower()
                    
                    # Store the whole row so we can extract data later
                    database[(brand, name)] = row
    except Exception as e:
        print(f"Failed to read dataset: {e}")
        
    return database

def main():
    print("Loading fragrance database... (This might take a few seconds for millions of rows)")
    database = load_dataset()
    if not database:
        print("Warning: No database loaded. Make sure you ran update_database.py first!")
        return
        
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb["Collection"]

        count = 0
        for r in range(2, sheet.max_row + 1):
            excel_brand = sheet.cell(row=r, column=2).value
            excel_name = sheet.cell(row=r, column=3).value
            
            if excel_brand and excel_name:
                search_brand = str(excel_brand).strip().lower()
                search_name = str(excel_name).strip().lower()
                
                key = (search_brand, search_name)
                if key in database:
                    row_data = database[key]
                    
                    # NOTE: These column extractions will need to be adjusted based on the EXACT 
                    # column headers in the Kaggle dataset you download.
                    # For now, it attempts to guess common column names.
                    
                    season = row_data.get('Season', row_data.get('season', "All Season"))
                    rating = row_data.get('Rating', row_data.get('rating', "8.0"))
                    longevity = row_data.get('Longevity', row_data.get('longevity', "Moderate"))
                    projection = row_data.get('Sillage', row_data.get('sillage', "Moderate"))
                    price = row_data.get('Price', row_data.get('price', "Budget"))
                    
                    # Write to excel
                    sheet.cell(row=r, column=4, value=season)       # Season
                    sheet.cell(row=r, column=5, value=longevity)    # Strength (guessing same as longevity)
                    sheet.cell(row=r, column=6, value=rating)       # Rating
                    sheet.cell(row=r, column=7, value=longevity)    # Longevity
                    sheet.cell(row=r, column=8, value=projection)   # Projection
                    sheet.cell(row=r, column=9, value=price)        # Price Tier
                    
                    count += 1

        wb.save(excel_path)
        print(f"Updated {count} fragrances successfully using the external database.")
    except Exception as e:
        print(f"Error updating excel: {e}")

if __name__ == "__main__":
    main()
