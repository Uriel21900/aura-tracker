import openpyxl
from copy import copy

excel_path = r"C:\Users\joseu\Downloads\Jose's Fragrance Collection Tracker.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    
    # Update data in primary tabs
    tabs_to_update = ["Collection", "Sorted by Season", "Sorted by Strength"]
    for tab in tabs_to_update:
        if tab in wb.sheetnames:
            sheet = wb[tab]
            for r in range(3, sheet.max_row + 1):
                brand = sheet.cell(row=r, column=2).value
                name = sheet.cell(row=r, column=3).value
                if brand and name:
                    brand_str = str(brand).strip().upper()
                    if brand_str == "KDJ" or brand_str == "KDJ INSPIRED":
                        sheet.cell(row=r, column=13).value = f"Dupe of {str(name).strip()}"

    # Recreate the "Sorted by Inspiration" tab to apply new sort order
    if "Sorted by Inspiration" in wb.sheetnames:
        del wb["Sorted by Inspiration"]
        
    source_ws = wb["Collection"]
    ws_new = wb.copy_worksheet(source_ws)
    ws_new.title = "Sorted by Inspiration"
    
    rows_data = []
    for r in range(3, ws_new.max_row + 1):
        if ws_new.cell(row=r, column=2).value or ws_new.cell(row=r, column=3).value:
            row_cells = []
            for c in range(1, 14):
                cell = ws_new.cell(row=r, column=c)
                cell_data = {
                    'value': cell.value,
                    'data_type': cell.data_type,
                    'font': copy(cell.font) if cell.has_style else None,
                    'border': copy(cell.border) if cell.has_style else None,
                    'fill': copy(cell.fill) if cell.has_style else None,
                    'number_format': copy(cell.number_format) if cell.has_style else None,
                    'protection': copy(cell.protection) if cell.has_style else None,
                    'alignment': copy(cell.alignment) if cell.has_style else None,
                }
                row_cells.append(cell_data)
            rows_data.append(row_cells)
            
    # Sort based on Column M (Index 12)
    def get_sort_key(row):
        val = row[12]['value'] 
        return str(val).lower() if val is not None else ""
        
    rows_data.sort(key=get_sort_key)
    
    # Write sorted data back
    for i, row_cells in enumerate(rows_data):
        r = i + 3
        for j, cell_data in enumerate(row_cells):
            c = j + 1
            cell = ws_new.cell(row=r, column=c)
            
            if c == 1:
                cell.value = i + 1
            elif c == 12: 
                cell.value = f"=COUNTIF('Wear Log'!B:B,C{r})"
            elif c == 11: 
                cell.value = f'=IF(COUNTIF(\'Wear Log\'!B:B,C{r})>0, MAXIFS(\'Wear Log\'!A:A, \'Wear Log\'!B:B, C{r}), "")'
            else:
                cell.value = cell_data['value']
                cell.data_type = cell_data['data_type']
                
            if cell_data['font']: cell.font = cell_data['font']
            if cell_data['border']: cell.border = cell_data['border']
            if cell_data['fill']: cell.fill = cell_data['fill']
            if cell_data['number_format']: cell.number_format = cell_data['number_format']
            if cell_data['protection']: cell.protection = cell_data['protection']
            if cell_data['alignment']: cell.alignment = cell_data['alignment']

    # Adjust width for Column M in all tabs
    for tab in tabs_to_update + ["Sorted by Inspiration"]:
        if tab in wb.sheetnames:
            sheet = wb[tab]
            max_length = 0
            for cell in sheet['M']:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions['M'].width = max_length + 2

    wb.save(excel_path)
    print("Successfully updated KDJ inspirations.")

except Exception as e:
    print(f"Error: {e}")
